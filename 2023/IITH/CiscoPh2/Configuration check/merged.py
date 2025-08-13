import os, json, re
import csv
from nornir import InitNornir
from nornir_napalm.plugins.tasks import napalm_cli
from nornir_utils.plugins.functions import print_result
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Clear the terminal
os.system("clear")

# Initialize Nornir
nr = InitNornir(config_file="config.yaml")

# Define the commands to be run
commands = [
    "show spanning-tree mst",
    "show version",
    "show running-config | i no ip http secure-server",
    "show running-config | i no ip http server",
    "show running-config | i transport input",
    "show running-config | i username",
    "show banner motd"
]

# Define the expected results for the HTTP commands
expected_results = [
    "no ip http secure-server",
    "no ip http server"
]

# Define the labels for the CSV output for the HTTP commands
command_labels = [
    "https-disabled ?",
    "http-disabled ?"
]

# Function to run the commands
def nornir_napalm_cli_commands_example(task):
    task.run(task=napalm_cli, commands=commands)

# Run the task
results = nr.run(task=nornir_napalm_cli_commands_example)
print_result(results)

# Extract VLAN mappings for each MST instance
mst_pattern = re.compile(r'##### MST(\d+)\s+vlans mapped:\s+(.+)')

# Extract SW Version, SW Image, and Model
sw_version_pattern = re.compile(r'SW Version\s+(\d+\.\d+\.\d+)')
sw_image_pattern = re.compile(r'SW Image\s+(\S+)')
model_pattern = re.compile(r'Model\s+(\S+)')

# Regex pattern to extract the username
username_pattern = re.compile(r'username (\S+)')

# Prepare CSV file
csv_filename = "combined_output.csv"
lines = []
lines.append(["Hostname", "SW Version", "SW Image", "Model", "MST Instance", "VLANs Mapped", "https-disabled ?", "http-disabled ?", "Transport Input", "Username", "Banner MOTD"])

# Dictionary to store combined results for each hostname
combined_results = {}

for hostname, result in results.items():
    for task_result in result:
        if task_result.result:
            output_mst = task_result.result[commands[0]]
            output_version = task_result.result[commands[1]]
            output_https = task_result.result[commands[2]]
            output_http = task_result.result[commands[3]]
            output_transport = task_result.result[commands[4]]
            output_username = task_result.result[commands[5]]
            output_banner = task_result.result[commands[6]]

            # Extract values from the table format at the end of the output
            table_pattern = re.compile(r'Switch Ports Model\s+SW Version\s+SW Image\s+Mode\s*\n'
                                       r'------ ----- -----              ----------        ----------            ----   \s*\n'
                                       r'\*    \d+ \d+    (\S+)\s+(\d+\.\d+\.\d+)\s+(\S+)\s+INSTALL', re.MULTILINE)

            table_match = table_pattern.search(output_version)

            if table_match:
                model = table_match.group(1)
                sw_version = table_match.group(2)
                sw_image = table_match.group(3)
            else:
                model = "N/A"
                sw_version = "N/A"
                sw_image = "N/A"

            # Extract MST instance and VLAN mappings
            mst_matches = mst_pattern.findall(output_mst)
            for match in mst_matches:
                mst_instance = match[0]
                vlans_mapped = match[1]

                # Ensure VLANs are separated by commas and treated as text
                vlans_mapped = ', '.join(vlans_mapped.split())
                vlans_mapped = f"'{vlans_mapped}"

                # Check HTTP and HTTPS command results
                https_disabled = "yes" if expected_results[0] in output_https else "no"
                http_disabled = "yes" if expected_results[1] in output_http else "no"

                # Check transport input result
                if "telnet" in output_transport or "all" in output_transport:
                    transport_input_status = "suspected telnet"
                else:
                    transport_input_status = "ssh only"

                # Find all matches of the username pattern
                matches = username_pattern.findall(output_username)
                usernames = ', '.join(matches)

                # Check banner MOTD result
                if output_banner.startswith("*****************"):
                    banner_motd_status = "yes"
                else:
                    banner_motd_status = "no"

                # Store combined results for each hostname
                if hostname not in combined_results:
                    combined_results[hostname] = {
                        "SW Version": sw_version,
                        "SW Image": sw_image,
                        "Model": model,
                        "https-disabled ?": https_disabled,
                        "http-disabled ?": http_disabled,
                        "Transport Input": transport_input_status,
                        "Username": usernames,
                        "Banner MOTD": banner_motd_status
                    }

                lines.append([
                    hostname,
                    combined_results[hostname]["SW Version"],
                    combined_results[hostname]["SW Image"],
                    combined_results[hostname]["Model"],
                    f"MST{mst_instance}",
                    vlans_mapped,
                    combined_results[hostname]["https-disabled ?"],
                    combined_results[hostname]["http-disabled ?"],
                    combined_results[hostname]["Transport Input"],
                    combined_results[hostname]["Username"],
                    combined_results[hostname]["Banner MOTD"]
                ])

# Write to CSV file
with open(csv_filename, mode='w', newline='') as file:
    writer = csv.writer(file)
    writer.writerows(lines)

# Create an Excel workbook and merge cells
excel_filename = "combined_output.xlsx"
wb = Workbook()
ws = wb.active

# Write the CSV data to the Excel sheet
with open(csv_filename, mode='r', newline='') as file:
    reader = csv.reader(file)
    for row in reader:
        ws.append(row)

# Merge cells for the same hostname
current_hostname = None
start_row = 2
for row in range(2, ws.max_row + 1):
    hostname = ws.cell(row=row, column=1).value
    if hostname != current_hostname:
        if current_hostname is not None:
            for col in range(1, 5):  # Merge columns 1 to 4 (Hostname, SW Version, SW Image, Model)
                ws.merge_cells(start_row=start_row, start_column=col, end_row=row - 1, end_column=col)
            for col in range(7, 12):  # Merge columns 7 to 11 (https-disabled ?, http-disabled ?, Transport Input, Username, Banner MOTD)
                ws.merge_cells(start_row=start_row, start_column=col, end_row=row - 1, end_column=col)
        current_hostname = hostname
        start_row = row

# Merge the last set of cells
for col in range(1, 5):  # Merge columns 1 to 4 (Hostname, SW Version, SW Image, Model)
    ws.merge_cells(start_row=start_row, start_column=col, end_row=ws.max_row, end_column=col)
for col in range(7, 12):  # Merge columns 7 to 11 (https-disabled ?, http-disabled ?, Transport Input, Username, Banner MOTD)
    ws.merge_cells(start_row=start_row, start_column=col, end_row=ws.max_row, end_column=col)

# Save the Excel file
wb.save(excel_filename)

print(f"Done. Check {csv_filename} and {excel_filename}")

import os
import re
from openpyxl import Workbook
from nornir import InitNornir
from nornir_napalm.plugins.tasks import napalm_cli
from nornir_utils.plugins.functions import print_result

# Clear the terminal
os.system("clear")

# Initialize Nornir
nr = InitNornir(config_file="config.yaml")

# Define the command to be run
command = "show spanning-tree blockedports"

# Function to run the command
def nornir_napalm_cli_commands_example(task):
    task.run(task=napalm_cli, commands=[command])

# Run the task
results = nr.run(task=nornir_napalm_cli_commands_example)
print_result(results)

# Prepare Excel file
filename = "blocked_ports_results.xlsx"
workbook = Workbook()
sheet = workbook.active
sheet.append(["Hostname", "MST Name", "Blocked Interface"])

# Regex pattern to extract MST names and blocked interfaces
blocked_ports_pattern = re.compile(r'(\S+)\s+(\S+)')

# Collect all results
all_results = []

# Process the results
for hostname, result in results.items():
    for task_result in result:
        if task_result.result:
            output = task_result.result[command]
            # Find all matches of the blocked ports pattern
            matches = blocked_ports_pattern.findall(output)
            valid_matches = []
            for match in matches:
                mst_name = match[0]
                blocked_interface = match[1]
                # Filter out unwanted text
                if mst_name not in ["Name", "Interfaces", "List", "Number", "of", "blocked", "ports", "(segments)", "in", "the", "system", ":", "0", "--------------------", "------------------------------------"]:
                    valid_matches.append([hostname, mst_name, blocked_interface])

            # If there are no valid matches, add a single row with N/A
            if not valid_matches:
                all_results.append([hostname, "N/A", "N/A"])
            else:
                all_results.extend(valid_matches)

# Sort the results by hostname
all_results.sort(key=lambda x: x[0])

# Write sorted results to Excel file
current_hostname = None
current_row = 2

for result in all_results:
    hostname, mst_name, blocked_interface = result
    if hostname != current_hostname:
        current_hostname = hostname
        current_row = sheet.max_row + 1
        sheet.append([hostname, mst_name, blocked_interface])
    else:
        sheet.append(["", mst_name, blocked_interface])

# Merge cells for the same hostname
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        if cell.value:
            start_row = cell.row
            end_row = start_row
            while sheet.cell(row=end_row + 1, column=1).value == cell.value:
                end_row += 1
            sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)

# Save the Excel file
workbook.save(filename)

print(f"Done. Check {filename}")
